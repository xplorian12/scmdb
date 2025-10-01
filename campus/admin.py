# campus/admin.py

from decimal import Decimal
import csv
from io import TextIOWrapper

from django import forms
from django.contrib import admin, messages
from django.db import transaction
from django.db.models import Count, F
from django.utils import timezone

from .models import School, Professor, Student, Roster
from .forms import RosterForm  # your CSV/XLSX upload fields etc.

PRICE_PER_STUDENT = Decimal("64.95")


# =========================
# Inlines
# =========================

class ProfessorInline(admin.TabularInline):
    model = Professor
    extra = 0
    fields = ("last_name", "first_name", "email", "department")
    show_change_link = True


class RosterInline(admin.TabularInline):
    """
    Shown on the Professor page to list that professor's rosters.
    Includes Created, Expiration, and computed Status.
    """
    model = Roster
    extra = 0
    fields = ("name", "created_at", "expiration_date", "status_inline", "source_filename")
    readonly_fields = ("created_at", "status_inline", "source_filename")
    show_change_link = True

    @admin.display(description="Status")
    def status_inline(self, obj):
        return "Expired" if getattr(obj, "is_expired", False) else "Active"


# --- Inline form adds a non-model "remove" checkbox we handle on save.
class RosterStudentInlineForm(forms.ModelForm):
    remove = forms.BooleanField(required=False, label="Remove")

    class Meta:
        model = Roster.students.through
        fields = "__all__"  # 'remove' is extra and rendered by the form


class RosterStudentInline(admin.TabularInline):
    """
    The ONLY place to manage students in THIS roster.
    - 'Remove' checkbox per row (we delete the membership in save_formset)
    - Shift-click range selection (requires static/admin/roster_inline_shift_select.js)
    - One blank row to add a student; filtered to this roster's professor
    """
    model = Roster.students.through
    form = RosterStudentInlineForm
    extra = 1
    can_delete = False  # we handle via 'remove'
    autocomplete_fields = ("student",)
    fields = ("remove", "student", "email")
    readonly_fields = ("email",)

    class Media:
        js = ("admin/roster_inline_shift_select.js",)

    def has_add_permission(self, request, obj):
        return True

    # Filter student choices to this roster's professor
    def get_formset(self, request, obj=None, **kwargs):
        self._parent_roster = obj
        return super().get_formset(request, obj, **kwargs)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "student":
            prof_id = getattr(getattr(self, "_parent_roster", None), "professor_id", None)
            if prof_id:
                kwargs["queryset"] = Student.objects.filter(professor_id=prof_id)
            else:
                kwargs["queryset"] = Student.objects.none()
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def email(self, obj):
        return getattr(obj.student, "email", "")


# =========================
# Admin action: increment purchased accounts on selected Schools
# =========================

class AddAccountsActionForm(forms.Form):
    delta = forms.IntegerField(min_value=1, label="Add accounts by")


@admin.action(description="Add to purchased accounts")
def add_purchased_accounts(modeladmin, request, queryset):
    raw = request.POST.get("delta")
    try:
        delta = int(raw)
    except (TypeError, ValueError):
        modeladmin.message_user(request, "Please enter a valid number in 'Add accounts by'.", level=messages.ERROR)
        return

    if delta <= 0:
        modeladmin.message_user(request, "Number must be greater than 0.", level=messages.ERROR)
        return

    updated = queryset.update(purchased_accounts=F("purchased_accounts") + delta)
    modeladmin.message_user(request, f"Added {delta} to {updated} school(s).", level=messages.SUCCESS)


# =========================
# Admin registrations
# =========================

@admin.register(School)
class SchoolAdmin(admin.ModelAdmin):
    """
    Clear hierarchy: Schools → Professors (inline).
    Shows Region (State or Country) and account counts.
    We annotate counts in the queryset to avoid per-row queries (more robust on SQLite).
    """
    list_display = (
        "name",
        "region_col",               # State or Country
        "purchased_accounts",
        "activated_accounts_col",   # from annotation, fallback to property
        "available_accounts_col",   # from annotation, fallback to property
        "created_at",
    )
    search_fields = ("name", "city")
    list_filter = ("state", "country")
    ordering = ("name",)
    inlines = [ProfessorInline]

    # --- NEW: annotate counts in one query (prevents per-row queries & sqlite lock issues)
    def get_queryset(self, request):
        from django.db.models import Count, F, Case, When, Value, IntegerField
        qs = super().get_queryset(request)
        qs = qs.annotate(
            activated_accounts_anno=Count("professors__rosters__students", distinct=True),
        ).annotate(
            available_accounts_anno=Case(
                When(purchased_accounts__gt=F("activated_accounts_anno"),
                     then=F("purchased_accounts") - F("activated_accounts_anno")),
                default=Value(0),
                output_field=IntegerField(),
            )
        )
        return qs

    @admin.display(description="Region")
    def region_col(self, obj):
        return obj.state or obj.country or ""

    @admin.display(description="Activated")
    def activated_accounts_col(self, obj):
        # prefer annotation; fall back to model property if not present
        val = getattr(obj, "activated_accounts_anno", None)
        if val is None:
            val = getattr(obj, "activated_accounts", 0)
        return val

    @admin.display(description="Available")
    def available_accounts_col(self, obj):
        # prefer annotation; fall back to model property if not present
        val = getattr(obj, "available_accounts_anno", None)
        if val is None:
            val = getattr(obj, "available_accounts", 0)
        return val


@admin.register(Student)
class StudentAdmin(admin.ModelAdmin):
    """
    Hidden from the sidebar, but registered so the green ➕ popup works.
    """
    search_fields = ("email", "first_name", "last_name", "professor__last_name", "professor__first_name")
    autocomplete_fields = ("professor",)

    def has_module_permission(self, request):
        return False


@admin.register(Professor)
class ProfessorAdmin(admin.ModelAdmin):
    """
    Middle of the hierarchy: School → Professor → Rosters inline.
    """
    list_display = ("last_name", "first_name", "school", "department", "email", "hire_date")
    search_fields = ("last_name", "first_name", "email", "department", "school__name")
    list_filter = ("school", "department")
    autocomplete_fields = ("school",)
    ordering = ("last_name", "first_name")
    inlines = [RosterInline]


@admin.register(Roster)
class RosterAdmin(admin.ModelAdmin):
    """
    Bottom object: has the student-membership inline only.
    Also shows Expiration + Status in list_display.
    """
    form = RosterForm
    change_form_template = "admin/campus/roster/change_form.html"  # your template (clipboard, etc.)

    # Do NOT show the M2M "students" field at the top; everything is in the inline.
    exclude = ("students",)

    list_display = (
        "status_col",
        "name",
        "professor",
        "school_col",
        "student_count_col",
        "discount_percent_col",
        "total_invoice_col",
        "expiration_date",   # requires models.py field
        "invoice_sent",
        "created_at",
    )
    list_display_links = ("name",)
    list_select_related = ("professor", "professor__school")
    search_fields = ("name", "professor__last_name", "professor__first_name", "professor__school__name")
    list_filter = ("professor__school", "invoice_sent")
    ordering = ("-created_at", "name")
    inlines = [RosterStudentInline]
    autocomplete_fields = ("professor",)

    # Remove the students field from the form even if ModelForm includes it
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        form.base_fields.pop("students", None)
        # Relabel & constrain the discount field as a percentage (0–100)
        if "discount_amount" in form.base_fields:
            f = form.base_fields["discount_amount"]
            f.label = "Discount (%)"
            f.help_text = "Enter a percent (0–100)."
            f.min_value = 0
            f.max_value = 100
            f.widget.attrs.setdefault("placeholder", "e.g., 15 for 15%")
        return form

    # Annotate student count for list page performance
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.annotate(student_count=Count("students", distinct=True))

    # Computed columns ---------------------------------

    @admin.display(description="School")
    def school_col(self, obj):
        return obj.professor.school

    @admin.display(description="Students")
    def student_count_col(self, obj):
        return getattr(obj, "student_count", obj.students.count())

    @admin.display(description="Discount %")
    def discount_percent_col(self, obj):
        val = obj.discount_amount or Decimal("0")
        return f"{val:.2f}%"

    @admin.display(description="Status")
    def status_col(self, obj):
        return "Expired" if getattr(obj, "is_expired", False) else "Active"

    @admin.display(description="Total invoice")
    def total_invoice_col(self, obj):
        """
        total = (count * price) * (1 - discount_percent/100)
        discount_amount is treated as a PERCENT (0–100).
        """
        count = getattr(obj, "student_count", obj.students.count())
        total_before = PRICE_PER_STUDENT * Decimal(count)
        pct = (obj.discount_amount or Decimal("0")) / Decimal("100")
        pct = max(Decimal("0"), min(Decimal("1"), pct))
        total = total_before * (Decimal("1") - pct)
        return f"${total:.2f}"

    # Pass emails to your change_form template for the “Copy to Clipboard” button
    def render_change_form(self, request, context, add=False, change=False, form_url="", obj=None):
        emails_csv = ""
        if obj:
            emails_csv = ", ".join(obj.students.values_list("email", flat=True))
        context["roster_emails_csv"] = emails_csv
        return super().render_change_form(request, context, add, change, form_url, obj)

    # Handle removal via the inline "remove" checkbox
    def save_formset(self, request, form, formset, change):
        # 1) Delete memberships marked for removal
        for f in formset.forms:
            if getattr(f, "cleaned_data", None):
                if f.cleaned_data.get("remove") and f.instance.pk:
                    f.instance.delete()
        # 2) Save remaining edits normally
        instances = formset.save(commit=False)
        for obj in instances:
            obj.save()
        formset.save_m2m()

    # Import (CSV/XLSX/XML) with robust column parsing (letter or number)
    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)

        upload = form.cleaned_data.get("csv_file")
        email_col_raw = (form.cleaned_data.get("email_column") or "").strip()
        skip_rows = form.cleaned_data.get("skip_rows") or 0

        if not upload:
            return  # no file provided

        def resolve_email_index(header_row, email_column_value):
            s = str(email_column_value or "").strip()
            if not s:
                raise ValueError("Email column is required.")

            # Number? (1-based)
            if s.isdigit():
                idx = int(s) - 1
                if idx < 0 or idx >= len(header_row):
                    raise ValueError("Email column number is out of range.")
                return idx

            # Letters? (A=1)
            if s.isalpha():
                def letters_to_index(v):
                    v = v.upper()
                    n = 0
                    for ch in v:
                        if not ('A' <= ch <= 'Z'):
                            raise ValueError(f"Bad column: {v}")
                        n = n * 26 + (ord(ch) - ord('A') + 1)
                    return n
                idx1 = letters_to_index(s)
                idx0 = idx1 - 1
                if 0 <= idx0 < len(header_row):
                    return idx0
                # otherwise fall through to header matching

            # Header name match (case-insensitive)
            wanted = s.lower()
            normalized = [str(h or "").strip().lower() for h in header_row]
            if wanted in normalized:
                return normalized.index(wanted)

            for c in ["email", "e-mail", "email address", "mail"]:
                if c in normalized:
                    return normalized.index(c)

            raise ValueError(
                f"Could not find email column '{email_column_value}'. "
                f"Available headers: {', '.join([str(x) for x in header_row])}"
            )

        def iter_rows_from_upload(upload_file):
            name = (getattr(upload_file, "name", "") or "").lower()

            # XLSX
            if name.endswith(".xlsx"):
                try:
                    from openpyxl import load_workbook
                except Exception:
                    messages.error(request, "To read .xlsx files, install openpyxl:  pip install openpyxl")
                    return None
                try:
                    upload_file.file.seek(0)
                    wb = load_workbook(upload_file.file, read_only=True, data_only=True)
                    ws = wb.active
                    for row in ws.iter_rows(values_only=True):
                        yield ["" if v is None else str(v) for v in row]
                    return
                except Exception as e:
                    messages.error(request, f"Couldn't read XLSX: {e}")
                    return None

            # Excel 2003 XML (SpreadsheetML)
            if name.endswith(".xml"):
                try:
                    import xml.etree.ElementTree as ET
                    upload_file.file.seek(0)
                    tree = ET.parse(upload_file.file)
                    root = tree.getroot()
                    ns = "{urn:schemas-microsoft-com:office:spreadsheet}"
                    rows = root.findall(f".//{ns}Row")
                    for r in rows:
                        values = []
                        for c in r.findall(f"{ns}Cell"):
                            d = c.find(f"{ns}Data")
                            values.append("" if d is None or d.text is None else str(d.text))
                        yield values
                    return
                except Exception as e:
                    messages.error(request, f"Couldn't parse Excel XML: {e}")
                    return None

            # Default: CSV with encoding fallbacks
            for enc in ("utf-8-sig", "cp1252", "latin-1"):
                try:
                    upload_file.file.seek(0)
                    wrapper = TextIOWrapper(upload_file.file, encoding=enc, newline="", errors="replace")
                    reader = csv.reader(wrapper)
                    for row in reader:
                        yield row
                    return
                except UnicodeDecodeError:
                    continue
            messages.error(request, "Couldn't decode the CSV. Please re-save as 'CSV UTF-8'.")
            return None

        rows_iter = iter_rows_from_upload(upload)
        if rows_iter is None:
            return

        # First row is header
        try:
            header = next(rows_iter)
        except StopIteration:
            messages.error(request, "The file appears to be empty.")
            return

        try:
            email_idx = resolve_email_index(header, email_col_raw)
        except ValueError as e:
            messages.error(request, str(e))
            return

        # Skip extra rows after header
        for _ in range(skip_rows):
            try:
                next(rows_iter)
            except StopIteration:
                break

        emails = []
        created_count = 0
        linked_count = 0
        professor = obj.professor

        with transaction.atomic():
            for row in rows_iter:
                if not row or email_idx >= len(row):
                    continue
                email = (row[email_idx] or "").strip().lower()
                if not email:
                    continue
                emails.append(email)

                student, created = Student.objects.get_or_create(
                    professor=professor,
                    email=email,
                    defaults={"first_name": "", "last_name": ""},
                )
                if created:
                    created_count += 1
                if not obj.students.filter(pk=student.pk).exists():
                    obj.students.add(student)
                    linked_count += 1

            # Save source filename once
            if hasattr(upload, "name") and not obj.source_filename:
                obj.source_filename = upload.name
                obj.save(update_fields=["source_filename"])

        messages.success(
            request,
            f"Processed {len(emails)} record(s). "
            f"New students: {created_count}. Linked to this roster: {linked_count}.",
        )
